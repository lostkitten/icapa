"""Atomic, multi-format report bundles for index research runs."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import asdict, replace
import math
import os
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

from icapa.analytics import AnalyticsRunResult, ResearchComparison
from icapa.workspace.identity import dataframe_content_digest

from .bundle_constants import (
    _EXCEL_MAX_ROWS,
    _HEADER_FILL,
    _HEADER_FONT,
    _PAYLOAD_FIELD_BY_SHEET,
    _V1_HEADER_ROWS,
)
from icapa.reporting.contracts import ReportBundleError, ReportPayload
from .security import (
    _flatten_mapping, _safe_cell, _sanitize_output_value,
    _sanitize_report_table,
)

def _collect_tables(
    payload: ReportPayload,
    analytics: AnalyticsRunResult | object | None,
    comparison: ResearchComparison | None,
    manifest: Mapping[str, Any] | list[Any] | object,
) -> dict[str, pd.DataFrame]:
    tables = {
        f"report_{name}": frame
        for name, frame in payload.sheet_frames().items()
    }
    if isinstance(analytics, AnalyticsRunResult):
        tables.update(analytics.tables())
    elif analytics is not None and callable(getattr(analytics, "tables", None)):
        for name, table in analytics.tables().items():
            if isinstance(table, pd.Series):
                table = table.rename("value").to_frame()
            if isinstance(table, pd.DataFrame):
                tables[f"analytics.{name}"] = table
    if comparison is not None:
        tables.update(
            {
                f"comparison.{name}": frame
                for name, frame in comparison.tables().items()
            }
        )
    manifest_rows = _flatten_mapping(manifest)
    tables["run_manifest"] = pd.DataFrame(
        manifest_rows,
        columns=["field", "value"],
    )
    return {
        name: _sanitize_report_table(frame)
        for name, frame in tables.items()
    }


def _summary_payload(
    payload: ReportPayload,
    analytics: AnalyticsRunResult | object | None,
    comparison: ResearchComparison | None,
    manifest: Any,
    *,
    contract_version: str,
) -> dict[str, Any]:
    result: dict[str, Any] = {
        "contract": "icapa-index-research-bundle",
        "contract_version": contract_version,
        "index_id": payload.index_id,
        "manifest": manifest,
    }
    if isinstance(analytics, AnalyticsRunResult):
        result["analytics_profile"] = analytics.spec.profile
        result["analytics"] = {
            plugin_id: {
                "metrics": _sanitize_output_value(plugin.metrics),
                "tables": sorted(plugin.tables),
            }
            for plugin_id, plugin in analytics.plugin_results.items()
        }
        result["analytics_diagnostics"] = [
            asdict(item) for item in analytics.diagnostics
        ]
    if comparison is not None:
        result["comparison"] = {
            "baseline": comparison.baseline_name,
            "candidates": list(comparison.candidate_names),
            "diagnostics": list(comparison.diagnostics),
        }
    return result


def _report_table_identity(
    name: str,
    frame: pd.DataFrame,
) -> dict[str, Any]:
    """Identify report values while excluding presentation-only wall time."""

    identity_frame = frame.copy(deep=True)
    if (
        name == "report_Overview"
        and {"field", "value"}.issubset(identity_frame.columns)
    ):
        generated = identity_frame["field"].eq("generated_at")
        identity_frame.loc[generated, "value"] = "<generated_at>"
    return {
        "schema": [
            (str(column), str(dtype))
            for column, dtype in zip(
                identity_frame.columns,
                identity_frame.dtypes,
                strict=True,
            )
        ],
        "content_digest": dataframe_content_digest(identity_frame),
    }


def _append_research_sheets(
    workbook_path: Path,
    analytics: AnalyticsRunResult | object | None,
    comparison: ResearchComparison | None,
    manifest: Any,
    *,
    v1_overflow: Sequence[tuple[str, pd.DataFrame]] = (),
) -> None:
    workbook = load_workbook(workbook_path, keep_links=True)
    if getattr(workbook, "_external_links", None):
        raise ReportBundleError("external workbook links are not permitted")
    additions: list[tuple[str, pd.DataFrame]] = list(v1_overflow)

    if isinstance(analytics, AnalyticsRunResult):
        metric_rows = []
        for plugin_id, plugin in analytics.plugin_results.items():
            for name, value in plugin.metrics.items():
                metric_rows.append(
                    {
                        "plugin": plugin_id,
                        "metric": name,
                        "value": value,
                    }
                )
        additions.append(
            (
                "Research Summary",
                pd.DataFrame(metric_rows, columns=["plugin", "metric", "value"]),
            )
        )
        table_map = {
            "calendar_period_performance.annual": "Calendar Periods",
            "rolling_risk.metrics": "Rolling Metrics",
            "constituent_change.detail": "Constituent Changes",
            "data_coverage.reviews": "Data Coverage",
        }
        all_tables = analytics.tables()
        for key, sheet_name in table_map.items():
            frame = all_tables.get(key)
            if frame is not None:
                additions.append((sheet_name, frame))
        # Explicit, normalized constraint diagnostics are authoritative.
        # Methodology diagnostics remain a compatibility fallback only when
        # the dedicated plugin did not produce a detail table.
        constraint_frame = all_tables.get("constraint_diagnostics.detail")
        if constraint_frame is None:
            constraint_frame = all_tables.get(
                "methodology_diagnostics.constraints"
            )
        if constraint_frame is not None:
            additions.append(
                ("Constraint Diagnostics", constraint_frame)
            )
    else:
        additions.append(
            (
                "Research Summary",
                pd.DataFrame(
                    [{"status": "Not available"}],
                    columns=["status"],
                ),
            )
        )

    additions.append(
        (
            "Run Manifest",
            pd.DataFrame(
                _flatten_mapping(manifest),
                columns=["field", "value"],
            ),
        )
    )
    if comparison is not None:
        additions.extend(
            (
                ("Comparison Summary", comparison.overview),
                ("Comparison Weights", comparison.weight_differences),
            )
        )

    for desired_name, frame in additions:
        _append_split_frames(workbook, desired_name, frame)
    temporary = workbook_path.with_name(f".{workbook_path.stem}.v2.xlsx")
    try:
        workbook.save(temporary)
        reopened = load_workbook(temporary, read_only=False, data_only=False)
        if getattr(reopened, "_external_links", None):
            raise ReportBundleError("external workbook links are not permitted")
        os.replace(temporary, workbook_path)
    finally:
        temporary.unlink(missing_ok=True)


def _split_v1_payload_for_v2(
    payload: ReportPayload,
) -> tuple[ReportPayload, tuple[tuple[str, pd.DataFrame], ...]]:
    """Keep the v1 renderer valid while preserving large v2 Excel tables.

    Excel v1 owns three non-data rows and remains unchanged. The v2 wrapper
    leaves the first valid chunk on the source template sheet and appends
    deterministic continuation sheets. Parquet tables continue to contain the
    complete unsplit frames.
    """

    maximum_rows = _EXCEL_MAX_ROWS - _V1_HEADER_ROWS
    if maximum_rows <= 0:
        raise ReportBundleError(
            "Excel row capacity must exceed the v1 template header rows"
    )
    replacements: dict[str, pd.DataFrame] = {}
    overflow: list[tuple[str, pd.DataFrame]] = []
    for sheet_name, field_name in _PAYLOAD_FIELD_BY_SHEET.items():
        frame = getattr(payload, field_name)
        if len(frame) <= maximum_rows:
            continue
        replacements[field_name] = frame.iloc[
            :maximum_rows
        ].reset_index(drop=True)
        remaining = frame.iloc[maximum_rows:].reset_index(drop=True)
        continuation_rows = _EXCEL_MAX_ROWS - 1
        for chunk_number, start in enumerate(
            range(0, len(remaining), continuation_rows),
            start=2,
        ):
            overflow.append(
                (
                    f"{sheet_name} {chunk_number}",
                    remaining.iloc[start : start + continuation_rows],
                )
            )
    return (
        payload if not replacements else replace(payload, **replacements),
        tuple(overflow),
    )


def _append_split_frames(workbook, name: str, frame: pd.DataFrame) -> None:
    frame = _sanitize_report_table(frame)
    safe_frame = frame.reset_index() if any(frame.index.names) else frame.copy()
    chunk_size = _EXCEL_MAX_ROWS - 1
    chunks = max(1, math.ceil(len(safe_frame) / chunk_size))
    for offset in range(chunks):
        suffix = "" if chunks == 1 else f" {offset + 1}"
        sheet_name = _unique_sheet_name(workbook, f"{name}{suffix}")
        sheet = workbook.create_sheet(sheet_name)
        chunk = safe_frame.iloc[
            offset * chunk_size : (offset + 1) * chunk_size
        ]
        if len(chunk.columns) == 0:
            sheet["A1"] = "status"
            sheet["A2"] = "Not available"
            continue
        for column_number, column_name in enumerate(chunk.columns, start=1):
            cell = sheet.cell(row=1, column=column_number)
            cell.value = _safe_cell(column_name)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        if chunk.empty:
            sheet["A2"] = "Not available"
        else:
            for row_number, record in enumerate(
                chunk.itertuples(index=False, name=None),
                start=2,
            ):
                for column_number, value in enumerate(record, start=1):
                    sheet.cell(
                        row=row_number,
                        column=column_number,
                    ).value = _safe_cell(value)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(chunk.columns))}{max(1, len(chunk) + 1)}"
        )


def _unique_sheet_name(workbook, requested: str) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", requested)[:31] or "Sheet"
    candidate = base
    sequence = 2
    while candidate in workbook.sheetnames:
        suffix = f" {sequence}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        sequence += 1
    return candidate



__all__ = [name for name in globals() if name.startswith("_")]
