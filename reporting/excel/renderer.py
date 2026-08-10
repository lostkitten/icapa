"""Render safe index research workbooks from the packaged ICAPA template."""

from __future__ import annotations

from datetime import date, datetime, timezone
import math
import os
from pathlib import Path
import re
import tempfile
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from ..builders import REPORT_CONTRACT, ReportPayload, SHEET_COLUMNS


TEMPLATE_VERSION = "1.0"
TEMPLATE_PATH = (
    Path(__file__).parent.parent
    / "templates"
    / "index_research_report.xlsx"
)
METADATA_SHEET = "_metadata"
NOT_AVAILABLE = "Not available"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_EXTERNAL_FORMULA = re.compile(
    r"(\[[^\]]+\]|https?://|file:|\\\\|(?:^|[\"'])/[A-Za-z]|[A-Za-z]:\\)",
    flags=re.IGNORECASE,
)


class ReportTemplateError(ValueError):
    """Raised when the packaged workbook violates the template contract."""


class ReportIntegrityError(ValueError):
    """Raised when a rendered workbook fails integrity validation."""


class ExcelReportRenderer:
    """Write a :class:`ReportPayload` into the fixed packaged workbook."""

    def render(
        self,
        payload: ReportPayload,
        output_path: str | os.PathLike[str],
        *,
        overwrite: bool = False,
    ) -> Path:
        """Render, atomically save, and reload-validate one report."""

        if not isinstance(payload, ReportPayload):
            raise TypeError("payload must be a ReportPayload")
        destination = Path(output_path)
        if destination.suffix.casefold() != ".xlsx":
            raise ReportIntegrityError("report output must use the .xlsx extension")
        if not destination.parent.exists():
            raise ReportIntegrityError("report output directory does not exist")
        if destination.exists() and not overwrite:
            raise FileExistsError(f"report already exists: {destination.name}")

        workbook = self._load_and_validate_template()
        for sheet_name, frame in payload.sheet_frames().items():
            self._write_frame(workbook[sheet_name], frame)
        metadata = workbook[METADATA_SHEET]
        metadata["A3"] = "index_id"
        metadata["B3"] = _safe_cell_value(payload.index_id)
        metadata.sheet_state = "hidden"
        self._validate_workbook_security(workbook)

        temporary = self._temporary_path(destination)
        try:
            workbook.save(temporary)
            self._validate_rendered_file(temporary, payload.index_id)
            if destination.exists() and not overwrite:
                raise FileExistsError(f"report already exists: {destination.name}")
            os.replace(temporary, destination)
        finally:
            temporary.unlink(missing_ok=True)
        return destination

    def _load_and_validate_template(self):
        if not TEMPLATE_PATH.is_file():
            raise ReportTemplateError("the packaged report template is missing")
        try:
            workbook = load_workbook(TEMPLATE_PATH, keep_links=True)
        except Exception as exc:
            raise ReportTemplateError("the packaged report template is unreadable") from exc
        self._validate_template_contract(workbook)
        self._validate_workbook_security(workbook)
        return workbook

    @staticmethod
    def _validate_template_contract(workbook) -> None:
        expected = [*SHEET_COLUMNS, METADATA_SHEET]
        if workbook.sheetnames != expected:
            raise ReportTemplateError(
                "report template sheet names or ordering do not match the contract"
            )
        metadata = workbook[METADATA_SHEET]
        if metadata["A1"].value != "template_version":
            raise ReportTemplateError("report template is missing its version marker")
        if str(metadata["B1"].value) != TEMPLATE_VERSION:
            raise ReportTemplateError(
                f"unsupported report template version: {metadata['B1'].value!r}"
            )
        if metadata["A2"].value != "report_contract":
            raise ReportTemplateError("report template is missing its contract marker")
        if metadata["B2"].value != REPORT_CONTRACT:
            raise ReportTemplateError("report template uses an unsupported contract")
        for sheet_name, columns in SHEET_COLUMNS.items():
            sheet = workbook[sheet_name]
            actual = tuple(
                sheet.cell(row=3, column=offset).value
                for offset in range(1, len(columns) + 1)
            )
            if actual != columns:
                raise ReportTemplateError(
                    f"{sheet_name} row 3 does not match the report column contract"
                )
            trailing = sheet.cell(row=3, column=len(columns) + 1).value
            if trailing not in (None, ""):
                raise ReportTemplateError(
                    f"{sheet_name} contains an unexpected row-3 column"
                )

    @staticmethod
    def _validate_workbook_security(workbook) -> None:
        if getattr(workbook, "_external_links", None):
            raise ReportTemplateError("external workbook links are not permitted")
        for defined_name in workbook.defined_names.values():
            expression = str(getattr(defined_name, "attr_text", "") or "")
            if _EXTERNAL_FORMULA.search(expression):
                raise ReportTemplateError(
                    "external references in defined names are not permitted"
                )
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if (
                        isinstance(value, str)
                        and value.startswith("=")
                        and _EXTERNAL_FORMULA.search(value)
                    ):
                        raise ReportTemplateError(
                            "external references in formulas are not permitted"
                        )

    @staticmethod
    def _write_frame(sheet, frame) -> None:
        columns = SHEET_COLUMNS[sheet.title]
        if tuple(frame.columns) != columns:
            raise ReportIntegrityError(
                f"{sheet.title} payload columns do not match the template"
            )
        _clear_output_region(sheet)
        sheet.freeze_panes = "A4"
        for column_number, column_name in enumerate(columns, start=1):
            cell = sheet.cell(row=3, column=column_number)
            cell.value = column_name
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if frame.empty:
            sheet["A4"] = NOT_AVAILABLE
            sheet["A4"].font = Font(italic=True, color="666666")
            sheet.auto_filter.ref = None
            _size_columns(sheet, columns)
            return

        for row_number, record in enumerate(
            frame.itertuples(index=False, name=None),
            start=4,
        ):
            for column_number, value in enumerate(record, start=1):
                cell = sheet.cell(row=row_number, column=column_number)
                cell.value = _safe_cell_value(value)
                cell.alignment = Alignment(vertical="top")
                _apply_number_format(cell, columns[column_number - 1])
        sheet.auto_filter.ref = (
            f"A3:{get_column_letter(len(columns))}{3 + len(frame)}"
        )
        _size_columns(sheet, columns)

    @classmethod
    def _validate_rendered_file(cls, path: Path, index_id: str) -> None:
        try:
            workbook = load_workbook(path, read_only=False, data_only=False)
        except Exception as exc:
            raise ReportIntegrityError("rendered report could not be reopened") from exc
        cls._validate_template_contract(workbook)
        cls._validate_workbook_security(workbook)
        metadata = workbook[METADATA_SHEET]
        if metadata.sheet_state != "hidden":
            raise ReportIntegrityError("rendered report metadata must be hidden")
        if metadata["A3"].value != "index_id" or metadata["B3"].value != index_id:
            raise ReportIntegrityError("rendered report index metadata is invalid")
        for sheet_name in SHEET_COLUMNS:
            if workbook[sheet_name]["A4"].value in (None, ""):
                raise ReportIntegrityError(
                    f"rendered report sheet is empty: {sheet_name}"
                )

    @staticmethod
    def _temporary_path(destination: Path) -> Path:
        handle = tempfile.NamedTemporaryFile(
            prefix=f".{destination.stem}.",
            suffix=".tmp.xlsx",
            dir=destination.parent,
            delete=False,
        )
        handle.close()
        return Path(handle.name)


def _clear_output_region(sheet) -> None:
    if sheet.max_row < 4:
        return
    for row in sheet.iter_rows(
        min_row=4,
        max_row=sheet.max_row,
        min_col=1,
        max_col=max(sheet.max_column, len(SHEET_COLUMNS[sheet.title])),
    ):
        for cell in row:
            cell.value = None


def _safe_cell_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        missing = pd.isna(value)
        if getattr(missing, "ndim", 0) == 0 and bool(missing):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "to_pydatetime"):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return value
    if hasattr(value, "item"):
        try:
            return _safe_cell_value(value.item())
        except (TypeError, ValueError):
            pass
    text = str(value)
    if len(text) > 32_767:
        raise ReportIntegrityError("Excel cell text exceeds 32,767 characters")
    if text.startswith(_FORMULA_PREFIXES):
        return f"'{text}"
    return text


def _apply_number_format(cell, column_name: str) -> None:
    if column_name.endswith("_date") or column_name in {"generated_at"}:
        cell.number_format = "yyyy-mm-dd"
    elif any(
        token in column_name
        for token in (
            "weight",
            "return",
            "turnover",
            "exposure",
            "contribution",
            "free_float",
        )
    ):
        cell.number_format = "0.000000%"
    elif column_name in {"price", "fx_rate", "market_cap", "shares", "value"}:
        cell.number_format = "0.000000"


def _size_columns(sheet, columns: tuple[str, ...]) -> None:
    for column_number, column_name in enumerate(columns, start=1):
        width = min(max(len(column_name) + 2, 12), 36)
        sheet.column_dimensions[get_column_letter(column_number)].width = width


__all__ = [
    "ExcelReportRenderer",
    "METADATA_SHEET",
    "NOT_AVAILABLE",
    "ReportIntegrityError",
    "ReportTemplateError",
    "TEMPLATE_PATH",
    "TEMPLATE_VERSION",
]
