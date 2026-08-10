"""Smoke tests for atomic multi-format index research report bundles."""

from dataclasses import dataclass
import json
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
import pytest

import icapa.reporting.bundle as bundle_module
import icapa.reporting.builders.bundle_tables as bundle_tables_module
from icapa.analytics import (
    AnalyticsDiagnostic,
    AnalyticsPluginResult,
    AnalyticsPluginSpec,
    AnalyticsRunResult,
    AnalyticsSpec,
    ResearchAnalyticsInputs,
    run_analytics_plugins,
)
from icapa.backtesting.reviews import BacktestResult
from icapa.reporting import (
    ReportBundleSpec,
    ReportDataError,
    ReportFormat,
    write_report_bundle,
)
from icapa.portfolio_construction.context import DataContext


def _runtime_sample(*parts: str) -> str:
    """Assemble redaction fixtures without storing sensitive-shaped literals."""
    return "".join(parts)


@dataclass
class _Workspace:
    workspace_name: str
    workspace_path: Path

    @property
    def reports_path(self):
        return self.workspace_path / "reports"


def _backtest():
    reviews = {}
    rows = []
    for effective_date, first_weight in (
        (pd.Timestamp("2026-03-23"), 0.6),
        (pd.Timestamp("2026-06-22"), 0.4),
    ):
        context = DataContext(
            reference_date=effective_date - pd.Timedelta(days=10),
            effective_date=effective_date,
            index_id="RESEARCH_INDEX",
        )
        context.set_dataframe(
            pd.DataFrame(
                [
                    {
                        "instrument_id": "A",
                        "name": "=not a formula",
                        "country": "US",
                        "industry": "Technology",
                        "benchmark_weight": 0.5,
                        "index_weight": first_weight,
                    },
                    {
                        "instrument_id": "B",
                        "name": "Example B",
                        "country": "GB",
                        "industry": "Industrials",
                        "benchmark_weight": 0.5,
                        "index_weight": 1.0 - first_weight,
                    },
                ]
            )
        )
        reviews[effective_date] = context
        frame = context.cons[["index_weight"]].reset_index()
        frame.insert(0, "effective_date", effective_date)
        rows.append(frame)
    return BacktestResult(
        weights=pd.concat(rows, ignore_index=True).set_index(
            ["effective_date", "instrument_id"]
        ),
        reviews=reviews,
    )


def test_report_bundle_writes_safe_excel_json_and_parquet(tmp_path):
    workspace = _Workspace("bundle-smoke", tmp_path / "workspace")
    backtest = _backtest()
    analytics = run_analytics_plugins(backtest)

    bundle = write_report_bundle(
        workspace,
        backtest,
        analytics=analytics,
        run_manifest={
            "definition_fingerprint": "a" * 64,
            "completed_at": "2026-07-27T12:00:00Z",
            "provider": {
                "name": "FactSet",
                "password": "must-not-appear",
            },
        },
        spec=ReportBundleSpec(name="smoke"),
    )

    assert bundle.path.is_dir()
    assert bundle.files["report.xlsx"].is_file()
    assert bundle.files["summary.json"].is_file()
    assert bundle.files["manifest.json"].is_file()
    assert any(name.endswith(".parquet") for name in bundle.files)
    workbook = load_workbook(bundle.files["report.xlsx"], data_only=False)
    assert "Research Summary" in workbook.sheetnames
    assert "Constituent Changes" in workbook.sheetnames
    assert "Run Manifest" in workbook.sheetnames
    assert workbook["Latest Holdings"]["B4"].value == "'=not a formula"
    manifest_text = bundle.files["manifest.json"].read_text(encoding="utf-8")
    assert "must-not-appear" not in manifest_text
    assert "[REDACTED]" in manifest_text
    checksums = json.loads(
        bundle.files["checksums.json"].read_text(encoding="utf-8")
    )
    assert checksums["report.xlsx"] == bundle.checksums["report.xlsx"]


def test_report_bundle_is_immutable_by_default(tmp_path):
    workspace = _Workspace("bundle-smoke", tmp_path / "workspace")
    backtest = _backtest()
    spec = ReportBundleSpec(name="immutable")

    write_report_bundle(workspace, backtest, spec=spec)

    try:
        write_report_bundle(workspace, backtest, spec=spec)
    except FileExistsError:
        pass
    else:
        raise AssertionError("the same report bundle must not be overwritten")


def test_report_bundle_overwrite_removes_replacement_backup(tmp_path):
    workspace = _Workspace("bundle-overwrite", tmp_path / "workspace")
    backtest = _backtest()
    first = write_report_bundle(
        workspace,
        backtest,
        spec=ReportBundleSpec(name="overwrite"),
    )

    replacement = write_report_bundle(
        workspace,
        backtest,
        spec=ReportBundleSpec(name="overwrite", overwrite=True),
    )

    assert replacement.path == first.path
    assert replacement.path.is_dir()
    assert not tuple(
        workspace.reports_path.glob(
            f".{first.bundle_id}.replaced-*"
        )
    )


def test_report_bundle_failed_overwrite_restores_original(
    tmp_path,
    monkeypatch,
):
    workspace = _Workspace("bundle-rollback", tmp_path / "workspace")
    backtest = _backtest()
    original = write_report_bundle(
        workspace,
        backtest,
        spec=ReportBundleSpec(name="rollback"),
    )
    original_files = {
        path.relative_to(original.path).as_posix(): path.read_bytes()
        for path in original.path.rglob("*")
        if path.is_file()
    }
    real_replace = bundle_module.os.replace
    publication_failed = False

    def fail_publication_once(source, destination):
        nonlocal publication_failed
        if Path(destination) == original.path and not publication_failed:
            publication_failed = True
            raise OSError("simulated publication failure")
        return real_replace(source, destination)

    monkeypatch.setattr(bundle_module.os, "replace", fail_publication_once)

    with pytest.raises(OSError, match="simulated publication failure"):
        write_report_bundle(
            workspace,
            backtest,
            spec=ReportBundleSpec(name="rollback", overwrite=True),
        )

    restored_files = {
        path.relative_to(original.path).as_posix(): path.read_bytes()
        for path in original.path.rglob("*")
        if path.is_file()
    }
    assert publication_failed
    assert restored_files == original_files
    assert not tuple(
        workspace.reports_path.glob(f".{original.bundle_id}.*")
    )


def test_report_bundle_identity_includes_table_values(tmp_path):
    workspace = _Workspace("bundle-content", tmp_path / "workspace")
    baseline = _backtest()
    candidate = _backtest()
    effective_date = min(candidate.reviews)
    candidate.reviews[effective_date].cons.loc["A", "index_weight"] = 0.7
    candidate.reviews[effective_date].cons.loc["B", "index_weight"] = 0.3
    candidate.weights.loc[
        (effective_date, "A"),
        "index_weight",
    ] = 0.7
    candidate.weights.loc[
        (effective_date, "B"),
        "index_weight",
    ] = 0.3
    spec = ReportBundleSpec(name="content")

    first = write_report_bundle(workspace, baseline, spec=spec)
    second = write_report_bundle(workspace, candidate, spec=spec)

    assert first.bundle_id != second.bundle_id
    assert first.path.is_dir()
    assert second.path.is_dir()


def test_sensitive_plugin_table_fields_are_redacted_in_every_format(tmp_path):
    workspace = _Workspace("bundle-redaction", tmp_path / "workspace")
    analytics = AnalyticsRunResult(
        spec=AnalyticsSpec(
            profile="sensitive_extension",
            plugins=(),
        ),
        plugin_results={
            "constraint_diagnostics": AnalyticsPluginResult(
                metrics={"api_token": "metrics-secret"},
                tables={
                    "detail": pd.DataFrame(
                        {
                            "password": [
                                "column-secret",
                                "column-secret",
                                "column-secret",
                            ],
                            "field": [
                                "ordinary_field",
                                "sql_query",
                                "ordinary_field",
                            ],
                            "value": [
                                "safe-value",
                                "query-secret",
                                "safe-value",
                            ],
                            "details": [
                                {"safe": "safe-value"},
                                {"safe": "safe-value"},
                                {"token": "nested-secret"},
                            ],
                        }
                    )
                },
            )
        },
        legacy_result=None,
        diagnostics=(),
    )

    bundle = write_report_bundle(
        workspace,
        _backtest(),
        analytics=analytics,
        run_manifest={
            "provider": {
                "password": "manifest-secret",
            }
        },
        methodology_parameters={
            "api_token": "methodology-secret",
            "serviceUri": "camel-uri-secret",
            "connectionUriValue": "camel-connection-secret",
            "account": "bundle-private-account",
            "file_path": "/private/bundle/input.csv",
            "authorization": "Bearer bundle-super-secret",
            "pem": _runtime_sample(
                "-----BEGIN PRIVATE ", "KEY-----\nbundle-private-key"
            ),
            "comment": "Basic YnVuZGxlLWJhc2ljLXNlY3JldA==",
            "location_hint": _runtime_sample(
                "/", "Users/private/bundle/source.csv"
            ),
            "accounting_method": "daily_accrual",
            "glide_path": "linear",
            "transition_path": "staged",
            "schema_version": "2026.1",
            "storage_settings": {
                "path": "/private/bundle/root",
                "database": "bundle-private-database",
                "schema": "bundle-private-schema",
                "server": "bundle-private-server",
                "warehouse": "bundle-private-warehouse",
                "oauth": "bundle-private-oauth",
                "glide_path": "nested-linear",
                "transition_path": "nested-staged",
                "schema_version": "2026.2",
            },
            "safe_parameter": 0.25,
        },
        spec=ReportBundleSpec(name="redaction"),
    )

    forbidden = {
        "api_token",
        "bundle-private-account",
        "bundle-private-database",
        "bundle-private-key",
        "bundle-private-oauth",
        "bundle-private-schema",
        "bundle-private-server",
        "bundle-private-warehouse",
        "bundle-super-secret",
        "ynvuzgxllwjhc2ljlxnly3jlda==",
        "camel-connection-secret",
        "camel-uri-secret",
        "column-secret",
        "connectionurivalue",
        "manifest-secret",
        "metrics-secret",
        "methodology-secret",
        "nested-secret",
        "password",
        "query-secret",
        "sql_query",
        "serviceuri",
        "token",
        "/private/bundle/input.csv",
        "/private/bundle/root",
        _runtime_sample("/", "users/private/bundle/source.csv"),
    }
    workbook = load_workbook(bundle.files["report.xlsx"], data_only=False)
    workbook_values = {
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    workbook_text = "\n".join(sorted(workbook_values)).casefold()
    assert not any(value in workbook_text for value in forbidden)
    assert "[redacted]" in workbook_text
    assert "accounting_method" in workbook_text
    assert "daily_accrual" in workbook_text
    assert "glide_path" in workbook_text
    assert "linear" in workbook_text
    assert "transition_path" in workbook_text
    assert "staged" in workbook_text
    assert "schema_version" in workbook_text
    assert "2026.1" in workbook_text
    assert "nested-linear" in workbook_text
    assert "nested-staged" in workbook_text
    assert "2026.2" in workbook_text
    methodology_sheet = workbook["Methodology Parameters"]
    methodology_keys = {
        str(row[0])
        for row in methodology_sheet.iter_rows(min_row=4, values_only=True)
        if row[0] is not None
    }
    assert "account" not in methodology_keys
    assert "file_path" not in methodology_keys
    for private_key in (
        "storage_settings.path",
        "storage_settings.database",
        "storage_settings.schema",
        "storage_settings.server",
        "storage_settings.warehouse",
        "storage_settings.oauth",
    ):
        assert private_key not in methodology_keys

    summary_text = bundle.files["summary.json"].read_text(
        encoding="utf-8"
    ).casefold()
    manifest_text = bundle.files["manifest.json"].read_text(
        encoding="utf-8"
    ).casefold()
    assert not any(value in summary_text for value in forbidden)
    assert not any(value in manifest_text for value in forbidden)
    assert "[redacted]" in summary_text
    assert "[redacted]" in manifest_text

    table_path = bundle.files[
        "tables/constraint_diagnostics.detail.parquet"
    ]
    restored = pd.read_parquet(table_path)
    parquet_text = restored.to_csv(index=True).casefold()
    assert not any(value in parquet_text for value in forbidden)
    assert "[redacted]" in parquet_text


def test_sensitive_string_values_are_redacted_in_every_report_format(tmp_path):
    workspace = _Workspace("string-redaction", tmp_path / "workspace")
    connection = (
        "postgresql://report_user:super-secret@db.internal:5432/research"
        "?sslmode=require&token=api-secret"
    )
    analytics = AnalyticsRunResult(
        spec=AnalyticsSpec(
            profile="generic_diagnostics",
            plugins=(),
        ),
        plugin_results={
            "generic_plugin": AnalyticsPluginResult(
                metrics={"calculation_note": connection},
                tables={
                    "detail": pd.DataFrame(
                        {"observation": [connection]}
                    )
                },
                diagnostics=(
                    AnalyticsDiagnostic(
                        level="warning",
                        code="provider_check",
                        message=connection,
                    ),
                ),
            )
        },
        legacy_result=None,
        diagnostics=(
            AnalyticsDiagnostic(
                level="warning",
                code="run_note",
                message=f"Provider response included {connection}",
            ),
        ),
    )

    bundle = write_report_bundle(
        workspace,
        _backtest(),
        analytics=analytics,
        run_manifest={"ordinary_note": connection},
        spec=ReportBundleSpec(name="string-redaction"),
    )

    forbidden = (
        "postgresql",
        "report_user",
        "super-secret",
        "db.internal",
        "api-secret",
    )
    json_text = "\n".join(
        bundle.files[name].read_text(encoding="utf-8").casefold()
        for name in ("summary.json", "manifest.json")
    )
    assert not any(token in json_text for token in forbidden)
    assert "[redacted]" in json_text

    workbook = load_workbook(bundle.files["report.xlsx"], data_only=False)
    workbook_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ).casefold()
    assert not any(token in workbook_text for token in forbidden)
    assert "[redacted]" in workbook_text

    parquet_text = "\n".join(
        pd.read_parquet(path).to_csv(index=True)
        for name, path in bundle.files.items()
        if name.endswith(".parquet")
    ).casefold()
    assert not any(token in parquet_text for token in forbidden)
    assert "[redacted]" in parquet_text


def test_colliding_normalized_parquet_table_names_remain_distinct(tmp_path):
    analytics = AnalyticsRunResult(
        spec=AnalyticsSpec(profile="filename_collision", plugins=()),
        plugin_results={
            "collision": AnalyticsPluginResult(
                tables={
                    "a/b": pd.DataFrame({"marker": ["slash"]}),
                    "a?b": pd.DataFrame({"marker": ["question"]}),
                    "plain": pd.DataFrame({"marker": ["plain"]}),
                }
            )
        },
        legacy_result=None,
        diagnostics=(),
    )

    first = write_report_bundle(
        _Workspace("collision-one", tmp_path / "workspace-one"),
        _backtest(),
        analytics=analytics,
        spec=ReportBundleSpec(
            name="collision-one",
            formats=(ReportFormat.PARQUET,),
        ),
    )
    second = write_report_bundle(
        _Workspace("collision-two", tmp_path / "workspace-two"),
        _backtest(),
        analytics=analytics,
        spec=ReportBundleSpec(
            name="collision-two",
            formats=(ReportFormat.PARQUET,),
        ),
    )

    first_names = {
        name for name in first.files if name.endswith(".parquet")
    }
    second_names = {
        name for name in second.files if name.endswith(".parquet")
    }
    assert first_names == second_names
    assert "tables/collision.plain.parquet" in first_names
    collision_names = sorted(
        name
        for name in first_names
        if name.startswith("tables/collision.a_b--")
    )
    assert len(collision_names) == 2
    assert collision_names[0] != collision_names[1]
    assert {
        pd.read_parquet(first.files[name])["marker"].iloc[0]
        for name in collision_names
    } == {"slash", "question"}


def test_sensitive_data_source_keys_are_hidden_before_contract_validation(
    tmp_path,
):
    workspace = _Workspace("source-redaction", tmp_path / "workspace")

    with pytest.raises(ReportDataError) as caught:
        write_report_bundle(
            workspace,
            _backtest(),
            data_sources={"sql_query": "data-source-secret"},
            spec=ReportBundleSpec(name="source-redaction"),
        )

    error_text = str(caught.value).casefold()
    assert "sql_query" not in error_text
    assert "data-source-secret" not in error_text
    assert "[redacted]" in error_text


def test_v2_constraint_sheet_prefers_explicit_normalized_diagnostics(
    tmp_path,
):
    workspace = _Workspace("constraint-report", tmp_path / "workspace")
    backtest = _backtest()
    for context in backtest.reviews.values():
        context.diagnostics["constraints"] = [
            {
                "name": "methodology_fallback",
                "value": 0.5,
                "upper": 0.7,
            }
        ]
    explicit = pd.DataFrame(
        [
            {
                "effective_date": "2026-03-23",
                "name": "explicit_maximum_weight",
                "value": 0.6,
                "lower": 0.0,
                "upper": 0.8,
            }
        ]
    )
    analytics = run_analytics_plugins(
        backtest,
        inputs=ResearchAnalyticsInputs(
            constraint_diagnostics=explicit,
        ),
    )

    bundle = write_report_bundle(
        workspace,
        backtest,
        analytics=analytics,
        spec=ReportBundleSpec(name="constraints"),
    )

    workbook = load_workbook(bundle.files["report.xlsx"], data_only=True)
    sheet = workbook["Constraint Diagnostics"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    constraint_name = headers.index("constraint_name")
    values = {row[constraint_name] for row in rows[1:]}
    assert values == {"explicit_maximum_weight"}


def test_v2_large_tables_split_into_deterministic_excel_sheets(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(bundle_tables_module, "_EXCEL_MAX_ROWS", 4)
    workspace = _Workspace("split-report", tmp_path / "workspace")
    backtest = _backtest()
    constraints = pd.DataFrame(
        [
            {
                "effective_date": "2026-03-23",
                "name": f"constraint_{position:02d}",
                "value": 0.1 * position,
                "lower": 0.0,
                "upper": 1.0,
            }
            for position in reversed(range(7))
        ]
    )
    analytics = run_analytics_plugins(
        backtest,
        spec=AnalyticsSpec(
            profile="constraint_only",
            plugins=(AnalyticsPluginSpec("constraint_diagnostics"),),
        ),
        inputs=ResearchAnalyticsInputs(
            constraint_diagnostics=constraints,
        ),
    )

    bundle = write_report_bundle(
        workspace,
        backtest,
        analytics=analytics,
        spec=ReportBundleSpec(
            name="split",
            formats=(ReportFormat.XLSX,),
        ),
    )

    workbook = load_workbook(bundle.files["report.xlsx"], data_only=True)
    sheet_names = [
        name
        for name in workbook.sheetnames
        if name.startswith("Constraint Diagnostics")
    ]
    assert sheet_names == [
        "Constraint Diagnostics 1",
        "Constraint Diagnostics 2",
        "Constraint Diagnostics 3",
    ]
    restored_names = []
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        name_position = list(rows[0]).index("constraint_name")
        restored_names.extend(
            row[name_position] for row in rows[1:] if row[name_position]
        )
    assert restored_names == [
        f"constraint_{position:02d}" for position in range(7)
    ]


def test_v2_splits_oversized_v1_tables_without_truncating_parquet(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(bundle_tables_module, "_EXCEL_MAX_ROWS", 6)
    workspace = _Workspace("base-table-split", tmp_path / "workspace")
    original = _backtest()
    rows = []
    for effective_date in original.reviews:
        for position in range(5):
            rows.append(
                {
                    "effective_date": effective_date,
                    "instrument_id": f"{effective_date:%Y%m%d}-{position:02d}",
                    "index_weight": 0.2,
                }
            )
    backtest = BacktestResult(
        weights=pd.DataFrame(rows).set_index(
            ["effective_date", "instrument_id"]
        ),
        reviews=original.reviews,
    )

    bundle = write_report_bundle(
        workspace,
        backtest,
        spec=ReportBundleSpec(
            name="base-split",
            formats=(ReportFormat.XLSX, ReportFormat.PARQUET),
        ),
    )

    workbook = load_workbook(bundle.files["report.xlsx"], data_only=True)
    assert [
        name
        for name in workbook.sheetnames
        if name.startswith("All Review Weights")
    ] == [
        "All Review Weights",
        "All Review Weights 2",
        "All Review Weights 3",
    ]
    restored_ids = []
    base_sheet = workbook["All Review Weights"]
    base_headers = [
        base_sheet.cell(row=3, column=column).value
        for column in range(1, 4)
    ]
    base_id_position = base_headers.index("instrument_id") + 1
    restored_ids.extend(
        value
        for row in range(4, base_sheet.max_row + 1)
        for value in (
            base_sheet.cell(row=row, column=base_id_position).value,
        )
        if value is not None
    )
    for sheet_name in ("All Review Weights 2", "All Review Weights 3"):
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        id_position = headers.index("instrument_id")
        restored_ids.extend(
            row[id_position]
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[id_position] is not None
        )

    parquet = pd.read_parquet(
        bundle.files[
            "tables/report_All_Review_Weights.parquet"
        ]
    )
    assert len(restored_ids) == len(parquet) == 10
    assert restored_ids == parquet["instrument_id"].tolist()
